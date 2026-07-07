# views.py
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from ..models import ProjectDF, ProjectAchatHardware, ProjectAchatSoftware, ProjectFinance,Notification_toDF
from ..serializers import (
    ProjectDFSerializer, ProjectDFDetailSerializer,
    ProjectAchatHardwareSerializer, ProjectAchatSoftwareSerializer,
    ProjectFinanceSerializer,
    Notification_toDFSerializer,
    NotificationToDFSerializerPlus
)
import pandas as pd
from django.http import HttpResponse
from rest_framework.decorators import action
from rest_framework.response import Response
from io import BytesIO
from rest_framework.views import APIView
from ..models import PERMISSIONS
from rest_framework.generics import RetrieveUpdateAPIView
from rest_framework.permissions import IsAuthenticated
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill
from django.db.models.functions import Upper, Trim, Coalesce
from django.db.models import Sum, F, Value, FloatField
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q, FloatField, IntegerField, DecimalField, DateField, DateTimeField
from datetime import date

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100

class ProjectDFViewSet(viewsets.ModelViewSet):
    queryset = ProjectDF.objects.all()
    serializer_class = ProjectDFSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['numero_appelle_doffre', 'client', 'secteur','constructeur','fournisseur',
                      'numero_contract','situation_projet', 'LOB','admin']
    filterset_fields = ['situation_projet', 'LOB', 'positionnement', 'previsionnel']
    ordering_fields = '__all__'  # ✅ allow ordering on all fields
    ordering = ['-id']
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = super().get_queryset()
        params = self.request.query_params
        user = self.request.user

        # --- 🧠 Permission-based filtering ---
        user_role_id = getattr(getattr(user, "Role", None), "id", None)

        if user_role_id is not None:
            if user_role_id not in PERMISSIONS.get("Commercial_View", []):
                queryset = queryset.filter(weightings_points__gt=29)
        # --- 📅 YEAR filter ---
        year_param = params.get("year")
        today = date.today()
        current_year = today.year

        if year_param and year_param.isdigit():
            year = int(year_param)
            queryset = queryset.filter(echeance__year=year)
        else:
            queryset = queryset.filter(
                Q(echeance__isnull=True) |
                Q(echeance__year=current_year) |
                Q(echeance__year=current_year - 1)
            )
        # Get visible fields from frontend
        visible_fields_param = params.get("visible_fields") if params.get("visible_fields") else ""
        visible_fields = (
            [f.strip() for f in visible_fields_param.split(",") if f.strip()]
            if visible_fields_param
            else None
        )

        # Define base field categories
        numeric_fields = [
            'montant_ht_marche', 'total_importation_equipement_usd', 
            'total_importation_service_usd', 'ps_usd', 'taux_change',
            'montant_travaux_realises', 'montant_travaux_restants', 'mt_achats',
            'marges_brutes', 'montant_dzd', 'montant_facture_dzd',
            'montant_encaisse_dzd', 'montant_travaux_factures_non_encaises_dzd',
            'montant_travaux_realises_non_factures_dzd','reste_a_realiser_dzd',
            'achats_hardware__mt_devises_cmd_a','achats_hardware__prix_dzd',
            'achats_hardware__dd','achats_hardware__tcs',
            'achats_hardware__tva','achats_hardware__taxes',
            'achats_software__mt_devises_brut','achats_software__mt_ne',
            'achats_software__mt_total_brut_dz','achats_software__mt_total_net_dza',
            'achats_software__dom_bancaire_4percent','achats_software__ibs',
            'achats_software__taxe','achats_software__dom_impot',
            'achats_software__tva','achats_software__predom_bancaire',
            'achats_software__dom_bancaire','achats_software__depot_dossier_dgi',
            'achats_software__accord_dgi','achats_software__transfert_bancaire',
            'finances__montent',
        ]
        date_fields = [
            'echeance','previsions_commandes', 'date_livraison_contractuelle',
            'date_cde','date_ods','date_po_contrat'
            'achats_software__date_depot_banque',
            'achats_software__date_previsionnelle_reception_atf',
            'achats_software__date_previsionnelle_transfert_fonds',
        ]
        search_fields = [
            'numero_appelle_doffre', 'client', 'secteur','constructeur','fournisseur',
            'numero_contract','situation_projet', 'LOB','admin'
            'achats_hardware__fournisseur','achats_hardware__numero_po',
            'achats_software__fournisseur','achats_software__numero_po',
            'achats_software__numero_facture','finances__num_facture_client',
        ]

        # If visible fields provided → filter only those
        if visible_fields:
            numeric_fields = [f for f in numeric_fields if f in visible_fields]
            date_fields = [f for f in date_fields if f in visible_fields]
            search_fields = [f for f in search_fields if f in visible_fields]

        # --- TEXT SEARCH
        search_text = params.get("searchtext")
        if search_text:
            q_obj = Q()
            for field in search_fields:
                q_obj |= Q(**{f"{field}__icontains": search_text})
            queryset = queryset.filter(q_obj)

        # --- NUMERIC RANGE
        search_min = params.get("searchmin")
        search_max = params.get("searchmax")
        if search_min or search_max:
            q_obj = Q()
            for field in numeric_fields:
                q_field = Q()
                if search_min:
                    q_field &= Q(**{f"{field}__gte": search_min})
                if search_max:
                    q_field &= Q(**{f"{field}__lte": search_max})
                q_obj |= q_field
            queryset = queryset.filter(q_obj)

        # --- DATE RANGE (skip Created_At)
        search_start_date = params.get("searchstartDate")
        search_end_date = params.get("searchendDate")
        if search_start_date or search_end_date:
            q_obj = Q()
            for field in date_fields:
                if field.lower() == "created_at":
                    continue
                q_field = Q()
                if search_start_date:
                    q_field &= Q(**{f"{field}__gte": search_start_date})
                if search_end_date:
                    q_field &= Q(**{f"{field}__lte": search_end_date})
                q_obj |= q_field
            queryset = queryset.filter(q_obj)

        return queryset

    def get_serializer_class(self):
        if self.action in ('list', 'retrieve'):
            return ProjectDFDetailSerializer
        return ProjectDFSerializer

    @action(detail=True, methods=['get'])
    def financial_summary(self, request, pk=None):
        project = self.get_object()
        summary = {
            'total_hardware': sum(hw.mt_devises_cmd_a for hw in project.achats_hardware.all()),
            'total_software': sum(sw.mt_devises_brut for sw in project.achats_software.all()),
            'total_finances': sum(finance.montent for finance in project.finances.all()),
            'total_paid': sum(finance.montent for finance in project.finances.filter(paid=True)),
            'total_unpaid': sum(finance.montent for finance in project.finances.filter(paid=False)),
        }
        return Response(summary)
    
    @action(detail=False, methods=['get'])
    def export_plan_de_charges(self, request):
        """Export PLAN DE CHARGES Excel file"""
        return self.export_to_excel(request, 'plan_de_charges')
    
    @action(detail=False, methods=['get'])
    def export_pipe_commercial(self, request):
        """Export PIPE COMMERCIAL Excel file"""
        return self.export_to_excel(request, 'pipe_commercial')
    
    @action(detail=False, methods=['get'])
    def export_import_hardware(self, request):
        """Export IMPORT HARDWARE Excel file"""
        return self.export_to_excel(request, 'import_hardware')
    
    @action(detail=False, methods=['get'])
    def export_import_software(self, request):
        """Export IMPORT SOFTWARE Excel file"""
        return self.export_to_excel(request, 'import_software')

    @action(detail=False, methods=['get'])
    def export_facture_client(self, request):
        """Export IMPORT SOFTWARE Excel file"""
        return self.export_to_excel(request, 'facture_client')

    @action(detail=False, methods=['get'])
    def export_globale(self, request):
        """Export GLOBALE Excel file with all fields"""
        return self.export_to_excel(request, 'globale')

    def export_to_excel(self, request, export_type):
        try:
            projects = self.get_queryset()
            projects = projects.exclude(weightings_points=100)
            
            project_row_ranges = None
            if export_type == 'plan_de_charges':
                data = self.get_plan_de_charges_data(projects)
            elif export_type == 'pipe_commercial':
                data = self.get_pipe_commercial_data(projects)
            elif export_type == 'import_hardware':
                projects = projects.exclude(weightings_points__in=[0, 1, 10])
                data, project_row_ranges = self.get_import_hardware_data(projects, return_ranges=True)
            elif export_type == 'import_software':
                projects = projects.exclude(weightings_points__in=[0, 1, 10])
                data, project_row_ranges = self.get_import_software_data(projects, return_ranges=True)
            elif export_type == 'facture_client':
                projects = projects.exclude(weightings_points__in=[0, 1, 10])
                data, project_row_ranges = self.get_facture_client_data(projects, return_ranges=True)
            elif export_type == 'globale':
                data = self.get_globale_data(projects)
            else:
                return Response({"error": "Invalid export type"}, status=400)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                if not isinstance(data, pd.DataFrame):
                    data = pd.DataFrame(data)

                # start writing data a bit lower to make space for title rows
                data.to_excel(writer, sheet_name='Export', startrow=2, index=False)
                ws = writer.sheets['Export']

                # === Add Title Row ===
                title_map = {
                    'plan_de_charges': "Plan de Charges",
                    'pipe_commercial': "Pipe Commercial",
                    'import_hardware': "Importation Hardware",
                    'import_software': "Importation Software",
                    'facture_client': "Facture Client",
                }
                title_text = title_map.get(export_type, "Export Données")

                total_cols = len(data.columns)
                today_str = datetime.now().strftime('%Y-%m-%d')

                # A1: red bold
                ws['A1'] = "******" # app_name
                ws['A1'].font = Font(size=15, bold=True, color="FF0000")
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

                # middle title cell (centered)
                mid_col = 5
                mid_letter = get_column_letter(mid_col)
                ws.cell(row=1, column=mid_col, value=title_text)
                ws.cell(row=1, column=mid_col).font = Font(size=18, bold=True, color="000000")
                ws.cell(row=1, column=mid_col).alignment = Alignment(horizontal='center', vertical='center')

                # last cell: today’s date (right aligned)
                ws.cell(row=1, column=10, value=today_str)
                ws.cell(row=1, column=10).alignment = Alignment(horizontal='right', vertical='center')
                ws.cell(row=1, column=10).font = Font(size=11, color="333333")

                # add empty row before the title
                ws.insert_rows(1)

                # === Header styling ===
                header_row_index = 4
                header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                header_font = Font(bold=True)
                thin = Side(border_style="thin", color="000000")
                thin_border = Border(left=thin, right=thin)
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

                for col_idx in range(1, total_cols + 1):
                    cell = ws.cell(row=header_row_index, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    #cell.border = thin_border

                # === Merge & Format Cells for grouped rows ===
                if export_type in ['import_hardware', 'import_software', 'facture_client'] and project_row_ranges:
                    cols_to_merge = [
                        'Client', "N° Appel d’Offre", 'N° Contrat', 'Client/Affaire', 
                        'Admin', 'N° CONTRAT', 'DATE CDE', 'Date livraison contractuelle', 
                        'Date PO/CONTRAT FOURNI'
                    ]
                    ranges_list = (
                        list(project_row_ranges.values()) 
                        if isinstance(project_row_ranges, dict) 
                        else list(project_row_ranges) if isinstance(project_row_ranges, list) 
                        else []
                    )

                    for col_name in cols_to_merge:
                        if col_name not in data.columns:
                            continue
                        col_idx = data.columns.get_loc(col_name) + 1
                        col_letter = get_column_letter(col_idx)

                        for (start, end) in ranges_list:
                            if not start or not end:
                                continue
                            try:
                                start_row = int(start)
                                end_row = int(end)
                            except Exception:
                                continue
                            if end_row - start_row >= 1:
                                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                                top_cell = ws.cell(row=start_row, column=col_idx)
                                top_cell.alignment = center_align
                                top_cell.border = thin_border
                                top_cell.font = header_font
                            else:
                                single_cell = ws.cell(row=start_row, column=col_idx)
                                single_cell.alignment = center_align
                                single_cell.border = thin_border

                # === Style remaining cells ===
                max_row = ws.max_row
                max_col = ws.max_column
                for r in range(5, max_row + 1):
                    for c in range(1, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        if not cell.alignment or (cell.alignment and cell.alignment.wrap_text is None):
                            cell.alignment = Alignment(vertical='center', wrap_text=True)
                        cell.border = thin_border

                # === Adjust column widths ===
                for i, col in enumerate(data.columns, start=1):
                    column_letter = get_column_letter(i)
                    max_length = len(str(col)) + 2
                    for cell in ws[column_letter]:
                        val = cell.value
                        if val:
                            max_length = max(max_length, len(str(val)) + 2)
                    ws.column_dimensions[column_letter].width = min(max_length, 60)

            output.seek(0)
            filename = f"projects_{export_type}_{pd.Timestamp.now().strftime('%Y-%m-%d')}.xlsx"
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def get_plan_de_charges_data(self, projects):
        """Prepare data for PLAN DE CHARGES export"""
        data = []
        projects = projects.exclude(montant_travaux_restants__gt=0)
        for project in projects:
            row = {
                'STATUT du MARCHE': project.situation_projet,
                'CONTRACTANT / CLIENT': project.client,
                'DOMICILIATION BANCAIRE': project.domiciliation_bancaire or '',
                'NANTISSABLE Oui/Non': 'Oui' if project.nantissable else 'Non',
                'MONTANT DZD': project.montant_dzd or 0,
                'OBJET': project.objet or '',
                "DATE D'ODS": project.date_ods.strftime('%Y-%m-%d') if project.date_ods else '',
                "DELAIS D'EXECUTION": str(project.delais_execution) if project.delais_execution else '',
                'TAUX DE REALISATION %': project.taux_realisation_pourcentage or 0,
                'MONTANT FACTURE DZD': project.montant_facture_dzd or 0,
                'MONTANT ENCAISSE DZD': project.montant_encaisse_dzd or 0,
                'MONTANT DES TRAVAUX FACTURES NON ENCAISSES DZD': project.montant_travaux_factures_non_encaises_dzd or 0,
                'MONTANT DES TRAVAUX REALISES NON FACTURES DZD': project.montant_travaux_realises_non_factures_dzd or 0,
                'RESTE A REALISER DZD': project.reste_a_realiser_dzd or 0,
            }
            data.append(row)

        df = pd.DataFrame(data)

        if not df.empty:
            # Ensure numeric
            numeric_cols = [
                'MONTANT DZD', 'MONTANT FACTURE DZD', 'MONTANT ENCAISSE DZD',
                'MONTANT DES TRAVAUX FACTURES NON ENCAISSES DZD',
                'MONTANT DES TRAVAUX REALISES NON FACTURES DZD', 'RESTE A REALISER DZD',
                'TAUX DE REALISATION %'
            ]
            for col in numeric_cols:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

            # Totals
            totals_row = {
                'STATUT du MARCHE': 'TOTAUX',
                'CONTRACTANT / CLIENT': '',
                'DOMICILIATION BANCAIRE': '',
                'NANTISSABLE Oui/Non': '',
                'MONTANT DZD': df['MONTANT DZD'].sum(),
                'OBJET': '',
                "DATE D'ODS": '',
                "DELAIS D'EXECUTION": '',
                'TAUX DE REALISATION %': '',
                'MONTANT FACTURE DZD': df['MONTANT FACTURE DZD'].sum(),
                'MONTANT ENCAISSE DZD': df['MONTANT ENCAISSE DZD'].sum(),
                'MONTANT DES TRAVAUX FACTURES NON ENCAISSES DZD': df['MONTANT DES TRAVAUX FACTURES NON ENCAISSES DZD'].sum(),
                'MONTANT DES TRAVAUX REALISES NON FACTURES DZD': df['MONTANT DES TRAVAUX REALISES NON FACTURES DZD'].sum(),
                'RESTE A REALISER DZD': df['RESTE A REALISER DZD'].sum(),
            }
            # Create two empty rows
            blank_rows = pd.DataFrame([{c: '' for c in df.columns} for _ in range(2)])

            # CA calculations
            ca_1 = df['MONTANT FACTURE DZD'].sum() * 0.8
            ca_2 = df.loc[df['TAUX DE REALISATION %'] == 0, 'MONTANT DZD'].sum()

            current_year = datetime.now().year
            next_year = current_year + 1

            ca1_row = {key: '' for key in totals_row.keys()}
            ca1_row.update({
                'DOMICILIATION BANCAIRE': f'C.A  {current_year}  ',
                'NANTISSABLE Oui/Non': 'Prévisionnel :',
                'MONTANT DZD': ca_1,
            })

            ca2_row = {key: '' for key in totals_row.keys()}
            ca2_row.update({
                'DOMICILIATION BANCAIRE': f'C.A à reporter  ',
                'NANTISSABLE Oui/Non': f'sur {next_year} :',
                'MONTANT DZD': ca_2,
            })

            df = pd.concat([df, blank_rows, pd.DataFrame([totals_row]), blank_rows, pd.DataFrame([ca1_row, ca2_row])], ignore_index=True)
        
        return df

    def get_pipe_commercial_data(self, projects):
        """Prepare data for PIPE COMMERCIAL export"""
        data = []
        projects = projects.exclude(montant_travaux_restants__gt=0)
        for project in projects:
            row = {
                'Positionnement': 'Oui' if project.positionnement else 'Non',
                'Objet du marché/contrats': project.secteur or '',
                "num d'appele d'offre": project.numero_appelle_doffre or '',
                "Maitre d'ouvrage": project.client or '',
                'Intitulé générique': project.objet or '',
                'Montant HT du marché': project.montant_ht_marche or 0,
                'Total Importation Equipement $': project.total_importation_equipement_usd or 0,
                'previsionnale Oui / Non': 'Oui' if project.previsionnel else 'Non',
                'Total Importation Service $': project.total_importation_service_usd or 0,
                'PS $': project.ps_usd or 0,
                'Taux de change': project.taux_change or 0,
                'fournisseur': project.fournisseur or '',
                'constructeur': project.constructeur or '',
                'Délais de réalisation': str(project.delais_realisation) if project.delais_realisation else '',
                'Échéance': project.echeance.strftime('%Y-%m-%d') if project.echeance else '',
                'Prévisions Commandes': project.previsions_commandes.strftime('%Y-%m-%d') if project.previsions_commandes else '',
                'Montant des travaux réalisés': project.montant_travaux_realises or 0,
                'Montant des travaux restants': project.montant_travaux_restants or 0,
                'LOB': project.get_LOB_display() if project.LOB else '',
                'weightings points': f"{project.weightings_points}%",
                'Mt des achats': project.mt_achats or 0,
                'Marges brutes': project.marges_brutes or 0,
                '': project.comment or '',
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Add calculated totals
        if not df.empty:
            df['Montant HT du marché'] = df['Montant HT du marché'].fillna(0)
            df['Total Importation Equipement $'] = df['Total Importation Equipement $'].fillna(0)
            df['Total Importation Service $'] = df['Total Importation Service $'].fillna(0)
            df['PS $'] = df['PS $'].fillna(0)
            df['Montant des travaux réalisés'] = df['Montant des travaux réalisés'].fillna(0)
            df['Montant des travaux restants'] = df['Montant des travaux restants'].fillna(0)
            df['Mt des achats'] = df['Mt des achats'].fillna(0)
            df['Marges brutes'] = df['Marges brutes'].fillna(0)
            totals_row = {
                'Positionnement': 'TOTAUX',
                'Objet du marché/contrats': '',
                "num d'appele d'offre": '',
                "Maitre d'ouvrage": '',
                'Intitulé générique': '',
                'Montant HT du marché': df['Montant HT du marché'].sum(),
                'Total Importation Equipement $': df['Total Importation Equipement $'].sum(),
                'previsionnale Oui / Non': '',
                'Total Importation Service $': df['Total Importation Service $'].sum(),
                'PS $': df['PS $'].sum(),
                'Taux de change': '',
                'fournisseur': '',
                'constructeur': '',
                'Délais de réalisation': '',
                'Échéance': '',
                'Prévisions Commandes': '',
                'Montant des travaux réalisés': df['Montant des travaux réalisés'].sum(),
                'Montant des travaux restants': df['Montant des travaux restants'].sum(),
                'LOB': '',
                'weightings points': '',
                'Mt des achats': df['Mt des achats'].sum(),
                'Marges brutes': df['Marges brutes'].sum(),
                '': '',
            }
            df = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)
        
        return df
    
    def get_import_hardware_data(self, projects, return_ranges=False):
        """Prepare data for IMPORT HARDWARE export. Returns (df, ranges) when return_ranges=True.
        project_row_ranges is a list of (start_excel_row, end_excel_row) tuples.
        """
        data = []
        project_row_ranges = []
        # We write DataFrame with startrow=2 in export_to_excel, so:
        writer_startrow = 2
        header_row_excel = writer_startrow + 1  # header will be row 3
        first_data_row_excel = header_row_excel + 1  # data starts at row 4

        # optional filter used in your code
        projects = projects.filter(weightings_points__in=[30, 60])

        for project in projects:
            project_data_start_index = len(data)  # 0-based index in data list before adding this project's rows

            hardware_purchases = list(project.achats_hardware.all())

            if hardware_purchases:
                for purchase in hardware_purchases:
                    row = {
                        'Client/Affaire': project.numero_appelle_doffre or '',
                        'Admin': project.admin or '',
                        'N° CONTRAT': project.numero_contract or '',
                        'N° PO': purchase.numero_po or '',
                        'MT PO/CONTRAT HT': project.montant_ht_marche or 0,
                        'DATE CDE': project.date_cde.strftime('%Y-%m-%d') if project.date_cde else '',
                        'Date livraison contractuelle': project.date_livraison_contractuelle.strftime('%Y-%m-%d') if project.date_livraison_contractuelle else '',
                        'Date prev reception stock': purchase.date_prev_reception_stock or '',
                        'VOLET': purchase.volet or '',
                        'FOURNI': purchase.fournisseur or '',
                        'MODE DE PAIEMENT': purchase.get_mode_paiement_display() if purchase.mode_paiement else '',
                        'MT DEVISES CMD A': purchase.mt_devises_cmd_a or 0,
                        'STATUT': purchase.statut or '',
                        'PREVISION EXPEDITION': purchase.prevision_expidition or '',
                        'PRIX DZA': purchase.prix_dzd or 0,
                        'Taux DD':purchase.TauxDD or 0,
                        'DD': purchase.dd or 0,
                        'TCS': purchase.tcs or 0,
                        'TVA': purchase.tva or 0,
                        'TAXES': purchase.taxes or 0,
                    }
                    data.append(row)
            else:
                # one empty data row allocated to the project so merges still make sense
                data.append({
                    'Client/Affaire': project.numero_appelle_doffre or '',
                    'Admin': project.admin or '',
                    'N° CONTRAT': project.numero_contract or '',
                    'N° PO': '',
                    'MT PO/CONTRAT HT': project.montant_ht_marche or 0,
                    'DATE CDE': project.date_cde.strftime('%Y-%m-%d') if project.date_cde else '',
                    'Date livraison contractuelle': project.date_livraison_contractuelle.strftime('%Y-%m-%d') if project.date_livraison_contractuelle else '',
                    'Date prev reception stock': '',
                    'VOLET': '',
                    'FOURNI': '',
                    'MODE DE PAIEMENT': '',
                    'MT DEVISES CMD A': 0,
                    'STATUT': '',
                    'PREVISION EXPEDITION': '',
                    'PRIX DZA': 0,
                    'Taux DD': 0,
                    'DD': 0,
                    'TCS': 0,
                    'TVA': 0,
                    'TAXES': 0,
                })

            # compute number of rows added for this project
            project_data_end_index = len(data) - 1  # 0-based inclusive
            n_rows = project_data_end_index - project_data_start_index + 1

            # convert 0-based indices into excel (1-based) rows:
            start_excel_row = first_data_row_excel + project_data_start_index
            end_excel_row = first_data_row_excel + project_data_end_index

            project_row_ranges.append((start_excel_row, end_excel_row))

        df = pd.DataFrame(data)
        return (df, project_row_ranges) if return_ranges else df

    def get_import_software_data(self, projects, return_ranges=False):
        """Prepare data for IMPORT SOFTWARE export."""
        data = []
        project_row_ranges = []

        writer_startrow = 2
        header_row_excel = writer_startrow + 1
        first_data_row_excel = header_row_excel + 1

        projects = projects.filter(weightings_points__in=[30, 60])

        for project in projects:
            project_data_start_index = len(data)
            software_purchases = list(project.achats_software.all())

            if software_purchases:
                for purchase in software_purchases:
                    row = {
                        'Client/Affaire': project.numero_appelle_doffre or '',
                        'Admin': project.admin or '',
                        'N° CONTRAT': project.numero_contract or '',
                        'N° PO': purchase.numero_po or '',
                        'N° FACTURE A': purchase.numero_facture_a or '',
                        'Date PO/CONTRAT FOURNI': project.date_po_contrat.strftime('%Y-%m-%d') if project.date_po_contrat else '',
                        'STATUT': purchase.statut or '',
                        'DATE DEPOT BANQUE FRANSABANK': purchase.date_depot_banque.strftime('%Y-%m-%d') if purchase.date_depot_banque else '',
                        'date prévisionnelle reception de ATF': purchase.date_previsionnelle_reception_atf.strftime('%Y-%m-%d') if purchase.date_previsionnelle_reception_atf else '',
                        'Date Previsionnelle transfert des fonds': purchase.date_previsionnelle_transfert_fonds.strftime('%Y-%m-%d') if purchase.date_previsionnelle_transfert_fonds else '',
                        'N° FACTURE': purchase.numero_facture or '',
                        'MT DEVISES/BRUT': purchase.mt_devises_brut or 0,
                        'MT NET': purchase.mt_ne or 0,
                        'MT TOTAL BRUT DZD': purchase.mt_total_brut_dz or 0,
                        'TAUX TAXE': purchase.taux_taxe or 0,
                        'MT TOTAL NET DZA': purchase.mt_total_net_dza or 0,
                        'DOM BANCAIRE 4%': purchase.dom_bancaire_4percent or 0,
                        'IBS': purchase.ibs or 0,
                        'TAXE': purchase.taxe or 0,
                        'DOM IMPOT': purchase.dom_impot or 0,
                        'TVA': purchase.tva or 0,
                        'PREDOM BANCAIRE': purchase.predom_bancaire or 0,
                        'DOM BANCAIRE': purchase.dom_bancaire or 0,
                        'DEPOT DOSSIER DGI': purchase.depot_dossier_dgi or 0,
                        'ACCORD DGI': purchase.accord_dgi or 0,
                        'TRANSFERT BANCAIRE': purchase.transfert_bancaire or 0,
                        'OBSERVATION': purchase.statut or '',
                    }
                    data.append(row)
            else:
                data.append({
                    'Client/Affaire': project.numero_appelle_doffre or '',
                    'Admin': project.admin or '',
                    'N° CONTRAT': project.numero_contract or '',
                    'N° PO': '',
                    'N° FACTURE A': '',
                    'Date PO/CONTRAT FOURNI': project.date_po_contrat.strftime('%Y-%m-%d') if project.date_po_contrat else '',
                    'STATUT': '',
                    'DATE DEPOT BANQUE FRANSABANK': '',
                    'date prévisionnelle reception de ATF': '',
                    'Date Previsionnelle transfert des fonds': '',
                    'N° FACTURE': '',
                    'MT DEVISES/BRUT': 0,
                    'MT NET': 0,
                    'MT TOTAL BRUT DZD': 0,
                    'TAUX TAXE': 0,
                    'MT TOTAL NET DZA': 0,
                    'DOM BANCAIRE 4%': 0,
                    'IBS': 0,
                    'TAXE': 0,
                    'DOM IMPOT': 0,
                    'TVA': 0,
                    'PREDOM BANCAIRE': 0,
                    'DOM BANCAIRE': 0,
                    'DEPOT DOSSIER DGI': 0,
                    'ACCORD DGI': 0,
                    'TRANSFERT BANCAIRE': 0,
                    'OBSERVATION': '',
                })

            project_data_end_index = len(data) - 1
            start_excel_row = first_data_row_excel + project_data_start_index
            end_excel_row = first_data_row_excel + project_data_end_index
            project_row_ranges.append((start_excel_row, end_excel_row))

        df = pd.DataFrame(data)
        return (df, project_row_ranges) if return_ranges else df

    def get_facture_client_data(self, projects, return_ranges=False):
        """Prepare data for FACTURE CLIENT export"""
        data = []
        project_row_ranges = []

        writer_startrow = 2
        header_row_excel = writer_startrow + 1
        first_data_row_excel = header_row_excel + 1

        projects = projects.filter(weightings_points__in=[30, 60])

        for project in projects:
            project_data_start_index = len(data)
            finances = list(project.finances.all())

            if finances:
                for finance in finances:
                    data.append({
                        'Client': project.client or '',
                        'N° Appel d’Offre': project.numero_appelle_doffre or '',
                        'N° Contrat': project.numero_contract or '',
                        'N° Facture Client': (finance.num_facture_client or '').strip(),
                        'Montant': float(finance.montent or 0.0),
                        'Payé': 'Oui' if finance.paid else 'Non',
                    })
            else:
                data.append({
                    'Client': project.client or '',
                    'N° Appel d’Offre': project.numero_appelle_doffre or '',
                    'N° Contrat': project.numero_contract or '',
                    'N° Facture Client': '',
                    'Montant': 0.0,
                    'Payé': '',
                })

            project_data_end_index = len(data) - 1
            start_excel_row = first_data_row_excel + project_data_start_index
            end_excel_row = first_data_row_excel + project_data_end_index
            project_row_ranges.append((start_excel_row, end_excel_row))

        df = pd.DataFrame(data)
        return (df, project_row_ranges) if return_ranges else df

    def get_globale_data(self, projects):
        """Prepare unified GLOBALE export combining all data sources."""
        try:
            # Get each dataset
            plan_df = self.get_plan_de_charges_data(projects)
            plan_df['Source'] = 'Plan de Charges'

            pipe_df = self.get_pipe_commercial_data(projects)
            pipe_df['Source'] = 'Pipe Commercial'

            import_hw_df, _ = self.get_import_hardware_data(projects, return_ranges=True)
            import_hw_df['Source'] = 'Import Hardware'

            import_sw_df, _ = self.get_import_software_data(projects, return_ranges=True)
            import_sw_df['Source'] = 'Import Software'

            facture_df, _ = self.get_facture_client_data(projects, return_ranges=True)
            facture_df['Source'] = 'Facture Client'

            # Combine all into one
            combined_df = pd.concat(
                [plan_df, pipe_df, import_hw_df, import_sw_df, facture_df],
                ignore_index=True
            )

            # Optional: sort by Source or any other field (like project number)
            combined_df = combined_df.fillna('')

            return combined_df

        except Exception as e:
            print(f"Error combining global data: {e}")
            return pd.DataFrame()
    
    @action(detail=False, methods=['get'], url_path='export_constructeur_secteur_chart')
    def export_constructeur_secteur_chart(self, request):
        """Export Excel: Multi-bar chart (Constructeur × Secteur) with PS + Importations sums"""
        try:
            # 🔹 Get filtered projects
            projects = (
                self.get_queryset()
                .annotate(
                    constructeur_clean=Upper(Trim('constructeur')),
                    secteur_clean=Upper(Trim('secteur')),
                    total_sum=Coalesce(F('ps_usd'), Value(0.0)) +
                              Coalesce(F('total_importation_equipement_usd'), Value(0.0)) +
                              Coalesce(F('total_importation_service_usd'), Value(0.0))
                )
                .values('constructeur_clean', 'secteur_clean')
                .annotate(total_usd_sum=Sum('total_sum'))
            )

            if not projects.exists():
                return Response({"error": "Aucune donnée disponible pour ce graphique."}, status=400)

            # 🔹 Convert to DataFrame
            df = pd.DataFrame(list(projects))
            df = df.fillna({'constructeur_clean': 'INCONNU', 'secteur_clean': 'INCONNU'})
            df.rename(columns={
                'constructeur_clean': 'Constructeur',
                'secteur_clean': 'Secteur',
                'total_usd_sum': 'Total ($)'
            }, inplace=True)

            # 🔹 Pivot table: Constructeur × Secteur
            pivot = df.pivot_table(
                index="Constructeur",
                columns="Secteur",
                values="Total ($)",
                aggfunc="sum",
                fill_value=0
            )

            # 🔹 Add totals
            pivot["TOTAL"] = pivot.sum(axis=1)
            total_row = pivot.sum(axis=0).to_frame().T
            total_row.index = ["TOTAL"]
            pivot = pd.concat([pivot, total_row])

            # 🔹 Write to Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                pivot.to_excel(writer, sheet_name="Répartition")

            wb = load_workbook(output)
            ws = wb.active

            # Bold totals
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            for col in range(1, ws.max_column + 1):
                ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

            # 🔹 Create chart
            chart = BarChart()
            chart.title = "Répartition par Constructeur et Secteur (PS + Importations)"
            chart.y_axis.title = "Total ($)"
            chart.x_axis.title = "Constructeur"
            chart.type = "col"
            chart.style = 10
            chart.height = 10
            chart.width = 20
            chart.grouping = "clustered"

            max_row = ws.max_row - 1
            max_col = ws.max_column - 1
            cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
            data = Reference(ws, min_col=2, max_col=max_col, min_row=1, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            ws.add_chart(chart, f"A{ws.max_row + 3}")

            # 🔹 Dynamic filename
            month_year = datetime.now().strftime("%B_%Y")
            filename = f"Constructeur_Secteur_Chart_{month_year}.xlsx"

            # 🔹 Return Excel response
            final_output = BytesIO()
            wb.save(final_output)
            final_output.seek(0)

            response = HttpResponse(
                final_output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=500)
        
    @action(detail=False, methods=['get'], url_path='export_constructeur_pie_chart')
    def export_constructeur_pie_chart(self, request):
        try:
            # 🔹 Aggregate and normalize Constructeur (trim + uppercase)
            data_queryset = (
                self.get_queryset()
                .annotate(constructeur_clean=Upper(Trim('constructeur')))
                .values('constructeur_clean')
                .annotate(
                    ps_usd_sum=Sum(Coalesce(F('ps_usd'), Value(0.0, output_field=FloatField()))),
                    equip_usd_sum=Sum(Coalesce(F('total_importation_equipement_usd'), Value(0.0, output_field=FloatField()))),
                    service_usd_sum=Sum(Coalesce(F('total_importation_service_usd'), Value(0.0, output_field=FloatField()))),
                    montant_ht_sum=Sum(Coalesce(F('montant_ht_marche'), Value(0.0, output_field=FloatField())))
                )
                .order_by('constructeur_clean')
            )

            if not data_queryset.exists():
                return Response({"error": "Aucune donnée disponible pour ce graphique."}, status=400)

            # 🔹 Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Constructeur_Pie"

            # --- Title
            ws.merge_cells("A1:F1")
            ws["A1"] = "Répartition par Constructeur"
            ws["A1"].font = Font(size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")

            # --- Header
            headers = [
                "Constructeur",
                "$",
                "Total Équipement ($)",
                "Total Service ($)",
                "DZD"
            ]
            ws.append(headers)

            # --- Add data rows
            for item in data_queryset:
                constructeur = item["constructeur_clean"] or "INCONNU"
                ps_usd = float(item["ps_usd_sum"] or 0)
                equip = float(item["equip_usd_sum"] or 0)
                service = float(item["service_usd_sum"] or 0)
                montant = float(item["montant_ht_sum"] or 0)
                total_usd = ps_usd + equip + service
                ws.append([constructeur, total_usd, equip, service, montant])

            # --- Add total row
            total_total_usd = sum((item["ps_usd_sum"] or 0) + (item["equip_usd_sum"] or 0) + (item["service_usd_sum"] or 0) for item in data_queryset)
            total_equip = sum(item["equip_usd_sum"] or 0 for item in data_queryset)
            total_service = sum(item["service_usd_sum"] or 0 for item in data_queryset)
            total_montant = sum(item["montant_ht_sum"] or 0 for item in data_queryset)

            ws.append(["TOTAL", total_total_usd, total_equip, total_service, total_montant])
            for col in ["A", "B", "C", "D", "E"]:
                ws[f"{col}{ws.max_row}"].font = Font(bold=True)

            # --- Create Pie Chart (based on "Total ($)")
            chart = PieChart()
            chart.title = "Répartition des Constructeurs (Total USD)"

            # Define chart data and labels (exclude header & total)
            labels = Reference(ws, min_col=1, min_row=3, max_row=ws.max_row - 1)
            data = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row - 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(labels)

            # ✅ Show percentages and names
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showCatName = True
            chart.dataLabels.showVal = False
            chart.dataLabels.position = 'bestFit'

            # Chart styling
            chart.height = 12
            chart.width = 18
            ws.add_chart(chart, "G4")

            # --- Prepare Excel file response
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            current_month = datetime.now().strftime("%Y-%m")
            filename = f"Constructeur_Pie_Chart_{current_month}.xlsx"

            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            return Response({"error": str(e)}, status=500)
                
class ProjectAchatHardwareViewSet(viewsets.ModelViewSet):
    queryset = ProjectAchatHardware.objects.all()
    serializer_class = ProjectAchatHardwareSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['numero_po', 'fournisseur', 'statut']
    filterset_fields = ['mode_paiement', 'info_project']

class ProjectAchatSoftwareViewSet(viewsets.ModelViewSet):
    queryset = ProjectAchatSoftware.objects.all()
    serializer_class = ProjectAchatSoftwareSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['numero_po', 'fournisseur', 'statut', 'numero_facture']
    filterset_fields = ['info_project']

class ProjectFinanceViewSet(viewsets.ModelViewSet):
    queryset = ProjectFinance.objects.all()
    serializer_class = ProjectFinanceSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['num_facture_client']
    filterset_fields = ['paid', 'info_project']
    
    @action(detail=False, methods=['get'])
    def unpaid_invoices(self, request):
        unpaid = self.get_queryset().filter(paid=False)
        serializer = self.get_serializer(unpaid, many=True)
        return Response(serializer.data)
    
class PermissionViewDF(APIView):

    def get(self, request):

        return Response(PERMISSIONS)

class UserNotificationsView(APIView):
    def get(self, request):
        user = request.user
        notifs = Notification_toDF.objects.filter(To=user).select_related('Notification').order_by('Notification__Date_Time')
        serializer = NotificationToDFSerializerPlus(notifs, many=True)
        return Response(serializer.data)

class Notification_toDFRetrieveUpdateView(RetrieveUpdateAPIView):
    queryset = Notification_toDF.objects.all()
    serializer_class = Notification_toDFSerializer
    permission_classes = [IsAuthenticated]

    def perform_update(self, serializer):
        notif = self.get_object()
        if notif.To != self.request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("You do not have permission to update this notification.")
        serializer.save()   
